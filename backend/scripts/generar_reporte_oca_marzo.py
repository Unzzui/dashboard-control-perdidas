#!/usr/bin/env python3
"""Genera el reporte ejecutivo de efectividad de OCA para marzo 2026.

Lee backend/data/resultado_consolidado.parquet, filtra por Fecha ejecución
en marzo 2026, calcula las tres métricas de efectividad (Hallazgo, Operativa
y Ajustada por CGE) y escribe el archivo backend/data/reporte_oca_marzo_2026.xlsx
siguiendo la filosofía de Minimalismo Ejecutivo del proyecto.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# --- Paths ---
ROOT = Path(__file__).resolve().parents[2]
DATA_PATH = ROOT / "backend" / "data" / "resultado_consolidado.parquet"
OUTPUT_PATH = ROOT / "backend" / "data" / "reporte_oca_marzo_2026.xlsx"

# --- Periodo objetivo ---
YEAR = 2026
MONTH = 3  # marzo

# --- Paleta (CLAUDE.md - Minimalismo Ejecutivo) ---
OCA_BLUE = "294D6D"
SLATE_800 = "1E293B"
SLATE_500 = "64748B"
SLATE_400 = "94A3B8"
SLATE_200 = "E2E8F0"
SLATE_100 = "F1F5F9"
SLATE_50 = "F8FAFC"
WHITE = "FFFFFF"

# --- Etiquetas oficiales (campo Responsabilidad) ---
RESP_OCA = "Responsabilidad Contratista"
RESP_CGE = "Responsabilidad CGE"
RESP_SIN = "Sin asignar"  # etiqueta interna para campo vacío

# --- Resultados de visita ---
RV_NORMAL = "Normal"
RV_CNR = "CNR"
RV_FALLIDA = "Visita fallida"
RV_ANULACION = "Cierre por anulación"
RV_MANTENIMIENTO = "Mantenimiento Medidor"

# --- Estilos reusables ---
FONT_TITLE = Font(name="Calibri", size=16, bold=True, color=OCA_BLUE)
FONT_SECTION = Font(name="Calibri", size=10, bold=True, color=SLATE_500)
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color=WHITE)
FONT_ZONE = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_KPI_LABEL = Font(name="Calibri", size=9, color=SLATE_400)
FONT_KPI_VALUE = Font(name="Calibri", size=22, bold=True, color=SLATE_800)
FONT_KPI_HINT = Font(name="Calibri", size=9, color=SLATE_500)
FONT_CELL = Font(name="Calibri", size=10, color=SLATE_800)
FONT_CELL_DIM = Font(name="Calibri", size=10, color=SLATE_500)

FILL_HEADER = PatternFill("solid", fgColor=OCA_BLUE)
FILL_ZONE = PatternFill("solid", fgColor=SLATE_800)
FILL_ALT = PatternFill("solid", fgColor=SLATE_50)
FILL_KPI_CARD = PatternFill("solid", fgColor=WHITE)

THIN = Side(border_style="thin", color=SLATE_200)
BORDER_CELL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

FMT_INT = "#,##0"
FMT_PCT = "0.0%"


def write_table(
    ws: Worksheet,
    df: pd.DataFrame,
    start_row: int,
    start_col: int = 1,
    pct_cols: tuple[str, ...] = (),
    int_cols: tuple[str, ...] = (),
) -> int:
    """Escribe un DataFrame como tabla con estilos. Devuelve la fila siguiente libre.

    pct_cols: nombres de columnas a formatear como porcentaje.
    int_cols: nombres de columnas a formatear como entero con miles.
    Resto de columnas numéricas: formato general.
    """
    # Header
    for j, col in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + j, value=str(col).upper())
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_CELL
    # Filas
    for i, (_, row) in enumerate(df.iterrows()):
        excel_row = start_row + 1 + i
        is_alt = i % 2 == 1
        for j, col in enumerate(df.columns):
            val = row[col]
            cell = ws.cell(row=excel_row, column=start_col + j, value=val)
            cell.font = FONT_CELL
            cell.border = BORDER_CELL
            if is_alt:
                cell.fill = FILL_ALT
            if col in pct_cols:
                cell.number_format = FMT_PCT
                cell.alignment = ALIGN_RIGHT
            elif col in int_cols:
                cell.number_format = FMT_INT
                cell.alignment = ALIGN_RIGHT
            elif j == 0:
                cell.alignment = ALIGN_LEFT
            else:
                cell.alignment = ALIGN_RIGHT
    # Anchos de columna
    for j, col in enumerate(df.columns):
        letter = get_column_letter(start_col + j)
        max_len = max(
            len(str(col)),
            *[len(str(v)) for v in df[col].astype(str).tolist()],
        )
        ws.column_dimensions[letter].width = max(12, min(40, max_len + 4))
    return start_row + len(df) + 2


def write_title(ws: Worksheet, row: int, text: str, col: int = 1) -> int:
    """Escribe un título de hoja. Devuelve la siguiente fila libre."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = FONT_TITLE
    return row + 2


def write_section(ws: Worksheet, row: int, text: str, col: int = 1) -> int:
    """Escribe un encabezado de sección en mayúsculas grises."""
    cell = ws.cell(row=row, column=col, value=text.upper())
    cell.font = FONT_SECTION
    return row + 1


def load_data(path: Path) -> pd.DataFrame:
    """Carga el parquet consolidado."""
    return pd.read_parquet(path)


def filter_period(df: pd.DataFrame, year: int, month: int) -> pd.DataFrame:
    """Filtra registros cuya Fecha ejecución cae dentro del mes indicado.

    Las filas con Fecha ejecución nula o fuera del rango se descartan.
    """
    fechas = pd.to_datetime(df["Fecha ejecución"], format="ISO8601", errors="coerce")
    inicio = pd.Timestamp(year=year, month=month, day=1)
    fin = (inicio + pd.offsets.MonthBegin(1))
    mask = (fechas >= inicio) & (fechas < fin)
    return df.loc[mask].copy()


def compute_global_metrics(df: pd.DataFrame) -> dict:
    """Calcula KPIs globales del reporte.

    Returns dict con:
      total, normal, cnr, fallidas, anulacion, mantenimiento,
      fallidas_cge, fallidas_oca, fallidas_sin_asignar,
      kwh_cnr, pct_fallidas_cge,
      eff_hallazgo, eff_operativa, eff_ajustada
    """
    total = len(df)
    rv = df["Resultado visita"]
    resp = df["Responsabilidad"].fillna("")

    normal = int((rv == RV_NORMAL).sum())
    cnr = int((rv == RV_CNR).sum())
    fallidas = int((rv == RV_FALLIDA).sum())
    anulacion = int((rv == RV_ANULACION).sum())
    mantenimiento = int((rv == RV_MANTENIMIENTO).sum())

    is_fallida = rv == RV_FALLIDA
    fallidas_cge = int((is_fallida & (resp == RESP_CGE)).sum())
    fallidas_oca = int((is_fallida & (resp == RESP_OCA)).sum())
    fallidas_sin_asignar = fallidas - fallidas_cge - fallidas_oca

    kwh_cnr = int(df["kWh CNR"].sum())

    eff_hallazgo = (normal + cnr) / total if total else 0.0
    eff_operativa = 1 - (fallidas / total) if total else 0.0
    denom_ajustada = total - fallidas_cge
    eff_ajustada = (
        1 - ((fallidas - fallidas_cge) / denom_ajustada)
        if denom_ajustada
        else 0.0
    )
    pct_fallidas_cge = fallidas_cge / fallidas if fallidas else 0.0

    return {
        "total": total,
        "normal": normal,
        "cnr": cnr,
        "fallidas": fallidas,
        "anulacion": anulacion,
        "mantenimiento": mantenimiento,
        "fallidas_cge": fallidas_cge,
        "fallidas_oca": fallidas_oca,
        "fallidas_sin_asignar": fallidas_sin_asignar,
        "kwh_cnr": kwh_cnr,
        "pct_fallidas_cge": pct_fallidas_cge,
        "eff_hallazgo": eff_hallazgo,
        "eff_operativa": eff_operativa,
        "eff_ajustada": eff_ajustada,
    }


def compute_by_dimension(df: pd.DataFrame, dim_col: str) -> pd.DataFrame:
    """Agrupa por la dimensión dada y calcula métricas por grupo.

    Returns DataFrame con columnas:
      [dim_col], Total, Normal, CNR, Fallidas, Fallidas_CGE, Fallidas_OCA,
      Eff_Hallazgo, Eff_Operativa, Eff_Ajustada
    Filas vacías o nulas en la dimensión se etiquetan como "Sin asignar".
    """
    work = df.copy()
    work[dim_col] = work[dim_col].fillna("").replace("", "Sin asignar")

    rows = []
    for grupo, sub in work.groupby(dim_col, sort=True):
        m = compute_global_metrics(sub)
        rows.append({
            dim_col: grupo,
            "Total": m["total"],
            "Normal": m["normal"],
            "CNR": m["cnr"],
            "Fallidas": m["fallidas"],
            "Fallidas_CGE": m["fallidas_cge"],
            "Fallidas_OCA": m["fallidas_oca"],
            "Eff_Hallazgo": m["eff_hallazgo"],
            "Eff_Operativa": m["eff_operativa"],
            "Eff_Ajustada": m["eff_ajustada"],
        })
    return pd.DataFrame(rows)


def compute_failed_cross(df: pd.DataFrame) -> pd.DataFrame:
    """Cross-tab Resultado final × Responsabilidad para visitas fallidas.

    Returns DataFrame indexado por Resultado final con columnas:
      [RESP_OCA, RESP_CGE, "Sin asignar"]. Incluye totales en el orden dado.
    """
    fallidas = df[df["Resultado visita"] == RV_FALLIDA].copy()
    resp = fallidas["Responsabilidad"].fillna("").replace("", "Sin asignar")
    fallidas = fallidas.assign(_resp=resp)
    cross = pd.crosstab(fallidas["Resultado final"], fallidas["_resp"])
    # Garantizar las 3 columnas en orden estable
    for col in [RESP_OCA, RESP_CGE, "Sin asignar"]:
        if col not in cross.columns:
            cross[col] = 0
    cross = cross[[RESP_OCA, RESP_CGE, "Sin asignar"]]
    cross = cross.sort_values(by=cross.columns.tolist(), ascending=False)
    return cross


def build_metodologia_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Metodología")
    ws.sheet_view.showGridLines = False
    row = write_title(ws, 1, "Metodología del Reporte")

    bloques = [
        ("Fuente de datos",
         "Archivo backend/data/resultado_consolidado.parquet. "
         "Universo histórico: jul-2024 a abr-2026. Único contratista presente: OCA."),
        ("Periodo del reporte",
         "Marzo 2026. Filtro aplicado sobre el campo Fecha ejecución "
         "en el rango [2026-03-01, 2026-04-01). Las visitas sin Fecha "
         "ejecución se descartan del análisis."),
        ("Definiciones de Resultado de visita",
         "Normal: visita ejecutada, servicio sin anomalías. "
         "CNR: visita ejecutada, se detectó Consumo No Registrado (hallazgo positivo). "
         "Visita fallida: no se pudo completar la inspección. "
         "Cierre por anulación: orden anulada antes/durante ejecución. "
         "Mantenimiento Medidor: intervención técnica sobre el equipo."),
        ("Responsabilidad",
         "Campo Responsabilidad del registro. "
         "Responsabilidad Contratista: atribuible a OCA. "
         "Responsabilidad CGE: atribuible al cliente o a condiciones del sistema CGE "
         "(casa deshabitada, desconectado en BT/MT, sin empalme, sitio eriazo, "
         "zona peligrosa, sin acceso por caja tortuga, etc.). "
         "Sin asignar: el campo viene vacío en el origen y se reporta tal cual."),
        ("Métricas de efectividad",
         "Se reportan tres en paralelo:"),
        ("    Efectividad de Hallazgo",
         "(Normal + CNR) / Total visitas. Mide la proporción de visitas con resultado conclusivo."),
        ("    Efectividad Operativa",
         "1 - (Visitas fallidas / Total visitas). Mide la capacidad de OCA de completar la visita."),
        ("    Efectividad Ajustada por CGE",
         "1 - (Fallidas no CGE / (Total - Fallidas CGE)). Excluye del cálculo las "
         "visitas fallidas atribuibles a CGE. Refleja la efectividad real de OCA "
         "descontando las causas que no controla."),
        ("Alcance",
         "Solo se reporta volumen físico (kWh CNR detectados). "
         "No se incluye valorización monetaria ni comparativos contra otros meses."),
    ]

    for label, body in bloques:
        ws.cell(row=row, column=1, value=label).font = FONT_SECTION
        row += 1
        cell = ws.cell(row=row, column=1, value=body)
        cell.font = FONT_CELL
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 45
        row += 2

    ws.column_dimensions["A"].width = 110


def build_resultados_sheet(wb: Workbook, df: pd.DataFrame, metrics: dict) -> None:
    ws = wb.create_sheet("Resultados Globales")
    ws.sheet_view.showGridLines = False
    row = write_title(ws, 1, "Resultados Globales — Marzo 2026")
    row = write_section(ws, row, "Distribución por Resultado de Visita")
    row += 1

    total = metrics["total"]
    orden = [
        (RV_NORMAL, metrics["normal"]),
        (RV_CNR, metrics["cnr"]),
        (RV_FALLIDA, metrics["fallidas"]),
        (RV_ANULACION, metrics["anulacion"]),
        (RV_MANTENIMIENTO, metrics["mantenimiento"]),
    ]
    rows = [
        {"Resultado": k, "Cantidad": v, "% Total": (v / total) if total else 0.0}
        for k, v in orden
    ]
    rows.append({"Resultado": "TOTAL", "Cantidad": total, "% Total": 1.0})
    tabla = pd.DataFrame(rows)
    write_table(
        ws, tabla, start_row=row, start_col=1,
        pct_cols=("% Total",), int_cols=("Cantidad",),
    )


def build_fallidas_sheet(wb: Workbook, cross: pd.DataFrame) -> None:
    ws = wb.create_sheet("Fallidas y Responsabilidad")
    ws.sheet_view.showGridLines = False
    row = write_title(ws, 1, "Visitas Fallidas — Distribución por Responsabilidad")
    row = write_section(
        ws, row,
        "Causa raíz (Resultado final) cruzada con Responsabilidad",
    )
    row += 1

    tabla = cross.copy()
    tabla["Total"] = tabla.sum(axis=1)
    # Fila de totales
    total_row = pd.DataFrame(tabla.sum(axis=0)).T
    total_row.index = ["TOTAL"]
    tabla = pd.concat([tabla, total_row])
    tabla.insert(0, "Causa (Resultado final)", tabla.index)
    tabla = tabla.reset_index(drop=True)

    int_cols = (RESP_OCA, RESP_CGE, "Sin asignar", "Total")
    write_table(ws, tabla, start_row=row, int_cols=int_cols)


def build_efectividad_dimension_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    dim_label: str,
    dim_df: pd.DataFrame,
) -> None:
    """Construye una hoja con la tabla de efectividad por dimensión.

    dim_df viene de compute_by_dimension; se renombra la primera columna a dim_label.
    Se agrega una fila TOTAL al final con totales y efectividades recalculadas.
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    row = write_title(ws, 1, title)
    row = write_section(ws, row, f"Métricas por {dim_label}")
    row += 1

    df = dim_df.copy()
    first_col = df.columns[0]
    df = df.rename(columns={first_col: dim_label})

    # Recomputar efectividades del total a partir de los recuentos
    total = int(df["Total"].sum())
    fallidas = int(df["Fallidas"].sum())
    fallidas_cge = int(df["Fallidas_CGE"].sum())
    normal = int(df["Normal"].sum())
    cnr = int(df["CNR"].sum())
    eff_h = (normal + cnr) / total if total else 0.0
    eff_o = 1 - (fallidas / total) if total else 0.0
    denom = total - fallidas_cge
    eff_a = 1 - ((fallidas - fallidas_cge) / denom) if denom else 0.0
    total_row = {
        dim_label: "TOTAL",
        "Total": total,
        "Normal": normal,
        "CNR": cnr,
        "Fallidas": fallidas,
        "Fallidas_CGE": fallidas_cge,
        "Fallidas_OCA": int(df["Fallidas_OCA"].sum()),
        "Eff_Hallazgo": eff_h,
        "Eff_Operativa": eff_o,
        "Eff_Ajustada": eff_a,
    }
    df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

    # Renombrar columnas para presentación
    df = df.rename(columns={
        "Eff_Hallazgo": "Ef. Hallazgo",
        "Eff_Operativa": "Ef. Operativa",
        "Eff_Ajustada": "Ef. Ajustada CGE",
        "Fallidas_CGE": "Fallidas CGE",
        "Fallidas_OCA": "Fallidas OCA",
    })

    int_cols = ("Total", "Normal", "CNR", "Fallidas", "Fallidas CGE", "Fallidas OCA")
    pct_cols = ("Ef. Hallazgo", "Ef. Operativa", "Ef. Ajustada CGE")
    write_table(ws, df, start_row=row, int_cols=int_cols, pct_cols=pct_cols)


def build_resumen_sheet(wb: Workbook, m: dict) -> None:
    """Resumen ejecutivo con KPIs en bloques tipo card."""
    ws = wb.create_sheet("Resumen Ejecutivo", 0)
    ws.sheet_view.showGridLines = False

    write_title(ws, 1, "Reporte de Efectividad OCA — Marzo 2026")
    sub = ws.cell(row=2, column=1, value="Inspecciones ejecutadas en marzo 2026")
    sub.font = FONT_KPI_HINT

    # KPIs en una grilla 2x3 (filas 4-5 etiquetas/valores, fila 6 hint)
    # Cada KPI ocupa 2 columnas. Distribuimos 5 KPIs en filas 4-7 / 9-12 / 14-17.
    kpis = [
        ("Total Inspecciones",
         f"{m['total']:,}",
         "Visitas con Fecha ejecución en marzo 2026"),
        ("Resultado Conclusivo",
         f"{(m['normal'] + m['cnr']):,} ({m['eff_hallazgo']:.1%})",
         "Normal + CNR"),
        ("CNR Detectados",
         f"{m['cnr']:,} | {m['kwh_cnr']:,} kWh",
         "Casos y energía recuperada al sistema"),
        ("Visitas Fallidas",
         f"{m['fallidas']:,} ({m['pct_fallidas_cge']:.1%} CGE)",
         f"{m['fallidas_cge']:,} atribuibles a CGE"),
        ("Efectividad Operativa",
         f"{m['eff_operativa']:.1%}  →  {m['eff_ajustada']:.1%}",
         "Bruta vs Ajustada por responsabilidad CGE"),
    ]

    for i, (label, value, hint) in enumerate(kpis):
        # 2 KPIs por fila (cols 1-3 y 4-6); última fila puede tener uno solo
        block_row = 4 + (i // 2) * 5
        block_col = 1 + (i % 2) * 4
        l = ws.cell(row=block_row, column=block_col, value=label.upper())
        l.font = FONT_KPI_LABEL
        v = ws.cell(row=block_row + 1, column=block_col, value=value)
        v.font = FONT_KPI_VALUE
        h = ws.cell(row=block_row + 3, column=block_col, value=hint)
        h.font = FONT_KPI_HINT
        # Combinar etiquetas/valores en 3 columnas para ancho de card
        ws.merge_cells(start_row=block_row, start_column=block_col,
                       end_row=block_row, end_column=block_col + 2)
        ws.merge_cells(start_row=block_row + 1, start_column=block_col,
                       end_row=block_row + 1, end_column=block_col + 2)
        ws.merge_cells(start_row=block_row + 3, start_column=block_col,
                       end_row=block_row + 3, end_column=block_col + 2)

    # Mensaje de cierre defensivo
    cierre_row = 4 + ((len(kpis) + 1) // 2) * 5 + 2
    msg = (
        f"En marzo 2026 OCA ejecutó {m['total']:,} inspecciones con efectividad "
        f"operativa de {m['eff_operativa']:.1%}. Al ajustar por las "
        f"{m['fallidas_cge']:,} visitas fallidas cuya responsabilidad recae en "
        f"CGE, la efectividad real de OCA asciende a {m['eff_ajustada']:.1%}, "
        f"con {m['cnr']:,} CNR detectados y {m['kwh_cnr']:,} kWh recuperados al sistema."
    )
    cell = ws.cell(row=cierre_row, column=1, value=msg)
    cell.font = FONT_CELL
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=cierre_row, start_column=1, end_row=cierre_row, end_column=7)
    ws.row_dimensions[cierre_row].height = 60

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 22


def main() -> None:
    """Orquesta carga, cómputo y render del reporte."""
    raise NotImplementedError("Implementado en tareas posteriores")


if __name__ == "__main__":
    main()
